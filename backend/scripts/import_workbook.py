from __future__ import annotations

import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.customer import Customer
from app.models.loan import Loan, LoanStatus
from app.models.repayment import Repayment
from app.models.user import User, UserRole
from app.services.loan_service import quantize_amount, refresh_overdue_status, set_initial_loan_values


def parse_decimal(value, default: str = "0") -> Decimal:
    if value is None or value == "":
        return Decimal(default)
    return Decimal(str(value).strip())


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    return datetime.fromisoformat(str(value)).date()


def parse_datetime(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def get_sheet_rows(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    if sheet.max_row < 2:
        return []

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def pick_admin_user(db):
    admin = db.query(User).filter(User.role == UserRole.admin).order_by(User.id.asc()).first()
    if not admin:
        raise RuntimeError("No admin user found. Create an admin user before importing.")
    return admin


def recalculate_loan(loan: Loan, db):
    repayments = (
        db.query(Repayment)
        .filter(Repayment.loan_id == loan.id)
        .order_by(Repayment.paid_at.asc(), Repayment.id.asc())
        .all()
    )
    total_paid = quantize_amount(sum((Decimal(repayment.amount) for repayment in repayments), Decimal("0.00")))
    loan.total_paid = total_paid
    loan.current_balance = quantize_amount(Decimal(loan.total_payable) - total_paid)

    if loan.current_balance <= 0:
        loan.current_balance = Decimal("0.00")
        loan.status = LoanStatus.closed
        loan.days_overdue = 0
    else:
        refresh_overdue_status(loan)


def main():
    if len(sys.argv) < 2:
        print("Usage: python -m scripts.import_workbook <path-to-xlsx>")
        raise SystemExit(1)

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        raise SystemExit(1)

    workbook = load_workbook(workbook_path, data_only=True)
    customer_rows = get_sheet_rows(workbook, "customers")
    loan_rows = get_sheet_rows(workbook, "loans")
    repayment_rows = get_sheet_rows(workbook, "repayments")

    db = SessionLocal()
    try:
        admin = pick_admin_user(db)
        customer_map: dict[str, Customer] = {}
        loan_map: dict[str, Loan] = {}

        for row in customer_rows:
            customer_code = normalize_text(row.get("customer_code"))
            full_name = normalize_text(row.get("full_name"))
            phone = normalize_text(row.get("phone"))
            if not customer_code or not full_name or not phone:
                raise ValueError(f"Customer row missing required fields: {row}")

            customer = Customer(
                full_name=full_name,
                phone=phone,
                email=normalize_text(row.get("email")),
                address=normalize_text(row.get("address")),
                national_id=normalize_text(row.get("national_id")),
            )
            db.add(customer)
            db.flush()
            customer_map[customer_code] = customer

        for row in loan_rows:
            loan_code = normalize_text(row.get("loan_code"))
            customer_code = normalize_text(row.get("customer_code"))
            disbursed_at = parse_date(row.get("disbursed_at"))
            if not loan_code or not customer_code or not disbursed_at:
                raise ValueError(f"Loan row missing required fields: {row}")

            customer = customer_map.get(customer_code)
            if not customer:
                raise ValueError(f"Loan references unknown customer_code: {customer_code}")

            loan = Loan(
                customer_id=customer.id,
                created_by=admin.id,
                loan_amount=parse_decimal(row.get("loan_amount")),
                interest_rate=Decimal("0.00"),
                monthly_interest_rate=parse_decimal(row.get("monthly_interest_rate")),
                service_charge_rate=parse_decimal(row.get("one_time_fee_rate")),
                stamp_duty_rate=Decimal("0.00"),
                tenure_months=int(row.get("tenure_months")),
                installment_amount=parse_decimal(row.get("installment_amount")) if row.get("installment_amount") not in (None, "") else Decimal("0.00"),
                total_payable=Decimal("0.00"),
                total_paid=Decimal("0.00"),
                current_balance=Decimal("0.00"),
                disbursed_at=disbursed_at,
                next_due_date=disbursed_at,
                status=LoanStatus.active,
                days_overdue=0,
            )
            set_initial_loan_values(loan)

            imported_status = normalize_text(row.get("status"))
            if imported_status in {status.value for status in LoanStatus}:
                loan.status = LoanStatus(imported_status)

            db.add(loan)
            db.flush()
            loan_map[loan_code] = loan

        for row in repayment_rows:
            loan_code = normalize_text(row.get("loan_code"))
            if not loan_code:
                raise ValueError(f"Repayment row missing loan_code: {row}")

            loan = loan_map.get(loan_code)
            if not loan:
                raise ValueError(f"Repayment references unknown loan_code: {loan_code}")

            repayment = Repayment(
                loan_id=loan.id,
                recorded_by=admin.id,
                amount=quantize_amount(parse_decimal(row.get("amount"))),
                method=normalize_text(row.get("method")) or "cash",
                note=normalize_text(row.get("note")),
                paid_at=parse_datetime(row.get("paid_at")) or datetime.now(),
            )
            db.add(repayment)
            db.flush()

        for loan in loan_map.values():
            recalculate_loan(loan, db)

        db.commit()
        print(
            f"Imported {len(customer_rows)} customers, {len(loan_rows)} loans, "
            f"and {len(repayment_rows)} repayments from {workbook_path.name}."
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
