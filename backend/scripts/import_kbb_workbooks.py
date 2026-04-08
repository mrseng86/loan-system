from __future__ import annotations

import re
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.customer import Customer
from app.models.loan import Loan, LoanStatus
from app.models.repayment import Repayment
from app.models.user import User, UserRole
from app.services.loan_service import quantize_amount, refresh_overdue_status, set_initial_loan_values


DATE_FORMATS = [
    "%d %B %Y",
    "%d %b %Y",
    "%d/%m/%y",
    "%d/%m/%Y",
    "%d-%m-%Y",
]


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value, default: str = "0") -> Decimal:
    if value is None or value == "":
        return Decimal(default)
    return Decimal(str(value).replace(",", "").strip())


def parse_date_text(value: str):
    text = normalize_text(value)
    if not text:
        return None

    text = text.replace("Loan Date:", "").strip()
    if "(" in text:
        text = text.split("(")[0].strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime_text(value: str):
    text = normalize_text(value)
    if not text:
        return None

    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def get_cell_value(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def find_row(ws, label: str):
    for row in range(1, ws.max_row + 1):
        value = normalize_text(get_cell_value(ws, row, 1))
        if value and label.lower() in value.lower():
            return row
    return None


def find_first_row(ws, labels: list[str]):
    for label in labels:
        row = find_row(ws, label)
        if row:
            return row
    return None


def extract_customer_info(ws):
    name_line = normalize_text(get_cell_value(ws, 3, 1)) or ""
    address_line = normalize_text(get_cell_value(ws, 5, 1)) or ""
    contact_line = normalize_text(get_cell_value(ws, 6, 1)) or ""
    loan_date_row = find_row(ws, "Loan Date")
    loan_date_line = normalize_text(get_cell_value(ws, loan_date_row, 1)) if loan_date_row else ""
    tenor_line = normalize_text(get_cell_value(ws, 10, 1)) or ""

    if name_line.lower().startswith("customer name:"):
        full_name = name_line.split(":", 1)[1].strip()
    else:
        full_name = name_line.split("/ IC")[0].strip()

    ic_match = re.search(r"IC\s*(?:NO)?\s*:\s*(.+)$", name_line, re.IGNORECASE)
    national_id = ic_match.group(1).strip() if ic_match else None

    address = address_line.replace("Address:", "").strip() if address_line.lower().startswith("address:") else address_line
    phone = contact_line.replace("Contact:", "").strip() if contact_line.lower().startswith("contact:") else contact_line

    tenor_match = re.search(r"(\d+)", tenor_line or "")
    tenure_months = int(tenor_match.group(1)) if tenor_match else 0

    return {
        "full_name": full_name,
        "national_id": national_id,
        "address": address,
        "phone": phone,
        "loan_date": parse_date_text(loan_date_line),
        "tenure_months": tenure_months,
    }


def extract_loan_info(ws):
    loan_amount_row = find_first_row(ws, ["Loan Amount", "Advance Amount", "Previous loan revised", "Restructure account"])
    total_interest_row = find_row(ws, "Total Interest")

    if not loan_amount_row or not total_interest_row:
        raise ValueError(f"Unable to locate summary rows in sheet {ws.title}")

    loan_amount_label = normalize_text(get_cell_value(ws, loan_amount_row, 1)) or ""
    if "previous loan revised" in loan_amount_label.lower():
        loan_amount = parse_decimal(get_cell_value(ws, loan_amount_row, 3))
        next_row_label = normalize_text(get_cell_value(ws, loan_amount_row + 1, 1)) or ""
        if "advance amount" in next_row_label.lower():
            loan_amount += parse_decimal(get_cell_value(ws, loan_amount_row + 1, 3))
    elif "restructure account" in loan_amount_label.lower():
        loan_amount = Decimal("0.00")
        row = loan_amount_row
        while row <= ws.max_row:
            row_label = normalize_text(get_cell_value(ws, row, 1)) or ""
            if "restructure account" in row_label.lower():
                loan_amount += parse_decimal(get_cell_value(ws, row, 3))
                row += 1
                continue
            break
    else:
        loan_amount = parse_decimal(get_cell_value(ws, loan_amount_row, 3))

    service_sum = Decimal("0.00")
    row = loan_amount_row + 1
    while row <= ws.max_row:
      label = normalize_text(get_cell_value(ws, row, 1))
      if not label:
          row += 1
          continue
      if "service chg" in label.lower():
          service_sum += parse_decimal(get_cell_value(ws, row, 3))
          row += 1
          continue
      break

    total_interest = parse_decimal(get_cell_value(ws, total_interest_row, 3))
    monthly_rate_label = normalize_text(get_cell_value(ws, total_interest_row, 4)) or ""
    monthly_interest_rate = Decimal("0.00")
    monthly_match = re.search(r"([\d.]+)\s*%", monthly_rate_label)
    if monthly_match:
        monthly_interest_rate = parse_decimal(monthly_match.group(1))

    schedule_header_row = find_row(ws, "Year")
    if not schedule_header_row:
        raise ValueError(f"Unable to find repayment table in sheet {ws.title}")

    installment_label_row = find_first_row(ws, ["INST. PER MTH", "LOAN PYMT PER MTH"])
    installment_amount = Decimal("0.00")
    if installment_label_row:
        installment_amount = parse_decimal(get_cell_value(ws, installment_label_row + 1, 4))

    headers = [normalize_text(get_cell_value(ws, schedule_header_row, col)) for col in range(1, 15)]
    if installment_amount <= 0:
        total_payment_col = None
        for idx, header in enumerate(headers, start=1):
            if header and header.lower() == "total payment":
                total_payment_col = idx
                break
        if total_payment_col:
            first_schedule_row = schedule_header_row + 1
            installment_amount = parse_decimal(get_cell_value(ws, first_schedule_row, total_payment_col))

    one_time_fee_rate = Decimal("0.00")
    if loan_amount > 0:
        one_time_fee_rate = quantize_amount((service_sum / loan_amount) * Decimal("100"))

    return {
        "loan_amount": quantize_amount(loan_amount),
        "service_sum": quantize_amount(service_sum),
        "total_interest": quantize_amount(total_interest),
        "monthly_interest_rate": monthly_interest_rate,
        "installment_amount": quantize_amount(installment_amount),
        "schedule_header_row": schedule_header_row,
        "schedule_headers": headers,
        "one_time_fee_rate": one_time_fee_rate,
    }


def parse_payment_entries(cell_value, default_amount: Decimal):
    if cell_value is None or cell_value == "":
        return []

    if hasattr(cell_value, "year") and hasattr(cell_value, "month") and hasattr(cell_value, "day"):
        return [(datetime.combine(cell_value, datetime.min.time()), default_amount)]

    text = str(cell_value).strip()
    if not text:
        return []

    compact_matches = re.findall(
        r"(\d{2}/\d{2}/\d{2,4})\s*-\s*rm\s*([0-9]+(?:\.[0-9]+)?)",
        text,
        flags=re.IGNORECASE,
    )
    if compact_matches:
        entries = []
        for date_text, amount_text in compact_matches:
            paid_at = parse_datetime_text(date_text)
            amount = parse_decimal(amount_text)
            if paid_at and amount > 0:
                entries.append((paid_at, amount))
        if entries:
            return entries

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) == 1 and not lines[0].lower().startswith("-rm"):
        paid_at = parse_datetime_text(lines[0])
        return [(paid_at, default_amount)] if paid_at else []

    entries = []
    current_date = None
    for line in lines:
        if line.lower().startswith("-rm"):
            amount = parse_decimal(line.replace("-rm", "").replace("-RM", ""))
            if current_date is not None:
                entries.append((current_date, amount))
            continue

        parsed_date = parse_datetime_text(line)
        if parsed_date:
            current_date = parsed_date

    return entries


def find_header_index(headers, label):
    for idx, header in enumerate(headers, start=1):
        if header and header.lower() == label.lower():
            return idx
    return None


def extract_repayments(ws, schedule_header_row: int, headers: list[str | None]):
    repayments = []
    payment_date_col = find_header_index(headers, "Payment Date")
    total_payment_col = find_header_index(headers, "Total Payment")
    period_col = find_header_index(headers, "Period")

    if not payment_date_col or not total_payment_col or not period_col:
        return repayments

    row = schedule_header_row + 1
    while row <= ws.max_row:
        period = get_cell_value(ws, row, period_col)
        total_payment = get_cell_value(ws, row, total_payment_col)
        if period in (None, "") and total_payment in (None, ""):
            row += 1
            if row > schedule_header_row + 10:
                break
            continue

        default_amount = quantize_amount(parse_decimal(total_payment))
        payment_cell = get_cell_value(ws, row, payment_date_col)
        repayments.extend(parse_payment_entries(payment_cell, default_amount))
        row += 1

    return repayments


def recalculate_loan(loan: Loan, db):
    repayments = (
        db.query(Repayment)
        .filter(Repayment.loan_id == loan.id)
        .order_by(Repayment.paid_at.asc(), Repayment.id.asc())
        .all()
    )
    total_paid = quantize_amount(sum((Decimal(repayment.amount) for repayment in repayments), Decimal("0.00")))
    loan.total_paid = total_paid
    loan.current_balance = quantize_amount(Decimal(loan.total_payable) - total_paid)

    if loan.current_balance <= 0:
        loan.current_balance = Decimal("0.00")
        loan.status = LoanStatus.closed
        loan.days_overdue = 0
    else:
        refresh_overdue_status(loan)


def get_admin_user(db):
    admin = db.query(User).filter(User.role == UserRole.admin).order_by(User.id.asc()).first()
    if not admin:
        raise RuntimeError("No admin user found.")
    return admin


def import_workbook(workbook_path: Path, db, admin: User):
    workbook = load_workbook(workbook_path, data_only=True)
    imported = {"customers": 0, "loans": 0, "repayments": 0, "skipped": 0}

    for sheet_name in workbook.sheetnames:
        try:
            ws = workbook[sheet_name]
            customer_info = extract_customer_info(ws)
            loan_info = extract_loan_info(ws)
            repayment_entries = extract_repayments(ws, loan_info["schedule_header_row"], loan_info["schedule_headers"])

            existing_customer = (
                db.query(Customer)
                .filter(Customer.full_name == customer_info["full_name"], Customer.phone == customer_info["phone"])
                .first()
            )

            if existing_customer:
                customer = existing_customer
            else:
                customer = Customer(
                    full_name=customer_info["full_name"],
                    phone=customer_info["phone"],
                    email=None,
                    address=customer_info["address"],
                    national_id=customer_info["national_id"],
                )
                db.add(customer)
                db.flush()
                imported["customers"] += 1

            loan = Loan(
                customer_id=customer.id,
                created_by=admin.id,
                loan_amount=loan_info["loan_amount"],
                interest_rate=Decimal("0.00"),
                monthly_interest_rate=loan_info["monthly_interest_rate"],
                service_charge_rate=loan_info["one_time_fee_rate"],
                stamp_duty_rate=Decimal("0.00"),
                tenure_months=customer_info["tenure_months"],
                installment_amount=loan_info["installment_amount"],
                total_payable=Decimal("0.00"),
                total_paid=Decimal("0.00"),
                current_balance=Decimal("0.00"),
                disbursed_at=customer_info["loan_date"],
                next_due_date=customer_info["loan_date"],
                status=LoanStatus.active,
                days_overdue=0,
            )
            set_initial_loan_values(loan)
            db.add(loan)
            db.flush()
            imported["loans"] += 1

            for paid_at, amount in repayment_entries:
                if not paid_at or amount <= 0:
                    continue
                db.add(
                    Repayment(
                        loan_id=loan.id,
                        recorded_by=admin.id,
                        amount=quantize_amount(amount),
                        method="cash",
                        note=f"Imported from {workbook_path.name} / {sheet_name}",
                        paid_at=paid_at,
                    )
                )
                imported["repayments"] += 1

            db.flush()
            recalculate_loan(loan, db)
        except Exception as exc:
            imported["skipped"] += 1
            print(f"Skipped sheet {sheet_name}: {exc}")

    return imported


def main():
    imports_dir = Path(r"C:\Users\user\Documents\Playground\loan-management-system\backend\imports")
    if len(sys.argv) > 1:
        workbook_paths = [Path(arg).resolve() for arg in sys.argv[1:]]
    else:
        workbook_paths = sorted(imports_dir.glob("*.xlsx"))

    if not workbook_paths:
        print("No .xlsx files found to import.")
        raise SystemExit(1)

    db = SessionLocal()
    try:
        admin = get_admin_user(db)
        grand_total = {"customers": 0, "loans": 0, "repayments": 0}

        for workbook_path in workbook_paths:
            result = import_workbook(workbook_path, db, admin)
            for key in grand_total:
                grand_total[key] += result[key]
            print(
                f"Imported {workbook_path.name}: "
                f"{result['customers']} customers, {result['loans']} loans, {result['repayments']} repayments, {result['skipped']} skipped"
            )

        db.commit()
        print(
            f"Done. Total imported: {grand_total['customers']} customers, "
            f"{grand_total['loans']} loans, {grand_total['repayments']} repayments."
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
