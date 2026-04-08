from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.customer import Customer
from app.models.loan import Loan
from app.models.repayment import Repayment
from app.models.user import User, UserRole
from app.services.loan_service import quantize_amount, refresh_overdue_status
from scripts.import_kbb_workbooks import (
    extract_customer_info,
    extract_loan_info,
    extract_repayments,
)


def get_admin_user(db):
    admin = db.query(User).filter(User.role == UserRole.admin).order_by(User.id.asc()).first()
    if not admin:
        raise RuntimeError("No admin user found.")
    return admin


def recalculate_loan(loan: Loan, db):
    repayments = (
        db.query(Repayment)
        .filter(Repayment.loan_id == loan.id)
        .order_by(Repayment.paid_at.asc(), Repayment.id.asc())
        .all()
    )
    total_paid = quantize_amount(sum((repayment.amount for repayment in repayments), start=Decimal("0.00")))
    loan.total_paid = total_paid
    loan.current_balance = quantize_amount(loan.total_payable - total_paid)
    if loan.current_balance <= 0:
        loan.current_balance = quantize_amount(Decimal("0.00"))
        from app.models.loan import LoanStatus
        loan.status = LoanStatus.closed
        loan.days_overdue = 0
    else:
        refresh_overdue_status(loan)


def find_customer(db, customer_info):
    customer = None
    if customer_info["national_id"]:
        customer = db.query(Customer).filter(Customer.national_id == customer_info["national_id"]).first()
    if not customer:
        customer = (
            db.query(Customer)
            .filter(Customer.full_name == customer_info["full_name"], Customer.phone == customer_info["phone"])
            .first()
        )
    return customer


def find_loan(db, customer: Customer, customer_info, loan_info):
    candidates = (
        db.query(Loan)
        .filter(
            Loan.customer_id == customer.id,
            Loan.disbursed_at == customer_info["loan_date"],
            Loan.tenure_months == customer_info["tenure_months"],
        )
        .all()
    )
    for loan in candidates:
        if quantize_amount(loan.loan_amount) == quantize_amount(loan_info["loan_amount"]):
            return loan
    return candidates[0] if candidates else None


def repayment_exists(db, loan_id: int, paid_at, amount):
    existing = (
        db.query(Repayment)
        .filter(
            Repayment.loan_id == loan_id,
            Repayment.paid_at == paid_at,
            Repayment.amount == quantize_amount(amount),
        )
        .first()
    )
    return existing is not None


def repair_workbook(db, workbook_path: Path):
    admin = get_admin_user(db)
    workbook = load_workbook(workbook_path, data_only=True)
    added = 0
    skipped_sheets = 0

    for sheet_name in workbook.sheetnames:
        try:
            ws = workbook[sheet_name]
            customer_info = extract_customer_info(ws)
            loan_info = extract_loan_info(ws)
            repayment_entries = extract_repayments(ws, loan_info["schedule_header_row"], loan_info["schedule_headers"])

            customer = find_customer(db, customer_info)
            if not customer:
                print(f"Skipped {sheet_name}: customer not found")
                skipped_sheets += 1
                continue

            loan = find_loan(db, customer, customer_info, loan_info)
            if not loan:
                print(f"Skipped {sheet_name}: loan not found")
                skipped_sheets += 1
                continue

            sheet_added = 0
            for paid_at, amount in repayment_entries:
                if not paid_at or amount <= 0:
                    continue
                if repayment_exists(db, loan.id, paid_at, amount):
                    continue

                db.add(
                    Repayment(
                        loan_id=loan.id,
                        recorded_by=admin.id,
                        amount=quantize_amount(amount),
                        method="cash",
                        note=f"Repair import from {workbook_path.name} / {sheet_name}",
                        paid_at=paid_at,
                    )
                )
                sheet_added += 1
                added += 1

            if sheet_added:
                db.flush()
                recalculate_loan(loan, db)
                print(f"Updated {sheet_name}: added {sheet_added} missing repayments")
        except Exception as exc:
            skipped_sheets += 1
            print(f"Skipped {sheet_name}: {exc}")

    return added, skipped_sheets


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/repair_kbb_repayments.py <workbook-path> [more-workbooks...]")
        raise SystemExit(1)

    workbook_paths = [Path(arg).resolve() for arg in sys.argv[1:]]

    db = SessionLocal()
    try:
        total_added = 0
        total_skipped = 0
        for workbook_path in workbook_paths:
            added, skipped = repair_workbook(db, workbook_path)
            total_added += added
            total_skipped += skipped
        db.commit()
        print(f"Done. Added {total_added} missing repayments. Skipped {total_skipped} sheets.")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
